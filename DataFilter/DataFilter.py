#__main file__

#========================================================================================================

# in-built imports:
from pathlib import Path
import threading
import time
import os
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
import openpyxl

# user file imports
import Global 
from Write import writeExcel
from FillVals import *
import Display

# globals
global t1
global t2
global drop_att 

#=========================================================================================================

# Processing Funcs:- [Phase-1]
# (data input, validation, conversion, filtering, elimination) 

# read dataset:
def getData(flag=True):
    
    load_file = openpyxl.load_workbook(Global.file_path )
    get_data = load_file.active
    
    row = get_data.max_row - 1
    col = get_data.max_column
    
    if( Global.read_data == False ):
        print_l("\n")
        print_l("*loading file...")
        print_l(">>>done")

        print_l("\n")
        print_l("*analyzing data...")
        print_l("dataset dimensions (initial):")
        print_l("data colums: [" + str(col) + "]")
        print_l("data records: [" + str(row) + "]")
        print_l(">>>done")
        
        print_l("\n")
        print_l("*checking primary validation...")

    
    # primary validation
    if( row < 50 or col < 2 ):
        print_l("\n")
        print_l("error: in primary validation!")
        print_l(">>>[InsuficientDataError]")
        Global.exit_flag = True
        return
    
    if(not Global.read_data):
        print_l(">>>(satistfied)")
        Global.read_data = True
    
    if(flag):
        print_l("\n")
        print_l("*reading Data...")
        
        # defining 2-d list of size dataset-1
        data =  [[] for i in range(row) ]

        row_count = 0
        for rows in get_data.iter_rows(2, row+1):
            for columns in range(col):
                val = rows[columns].value
                data[row_count].append(val)
            row_count += 1
        
        print_l(">>>done")
        
        return data
    
    else:
        # for attributes
        data = []
        print_l("\n")
        print_l("*reading attributes...")
        for rows in get_data.iter_rows(1,1):
            for columns in range(col):
                if(rows[columns].value == None):
                    data.append("    -    ")
                    continue
                
                val = rows[columns].value
                data.append(val)
        print_l(">>>done")
        return data 
    
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -   

# getting column information as data-type & missing value's count:
def get_column_info( data, rows, cols ):
    
    info_list = [ ["",0] for i in range(cols)]
    print_l("\n")
    print_l("column information: [data-type]  [missing value count]")
    
    sample_type = [1, "1"]
    for col in range(cols):
        
        type_count = [0,0]
        for row in range(rows):
            
            if(data[row][col] == None):
                info_list[col][1] += 1
                continue
            
            if( type(data[row][col]) == type(sample_type[1]) ):
                type_count[1] += 1
            else:
                type_count[0] += 1
            
        if( type_count.index( max(type_count) ) == 0):
            info_list[col][0] = "NUM"
        else:
            info_list[col][0] = "STR"
            
        print_l("column " + str(col) + ": [ " + info_list[col][0] + " ] " + "  [ " + str(info_list[col][1]) + " ]")
        
    return info_list  
    
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -   

# secondary validation:
def sec_validation( column_info_list ):
    
    print_l("\n")
    print_l("*checking secondary validation...")
    
    cols = len(column_info_list)
    label_col_idx = cols - 1
    
    if( column_info_list[label_col_idx][0] == "STR" ):
        print_l("error: can`t find label column")
        print_l(">>>[IncompleteDataError]")
        Global.exit_flag = True
        return 
    
    col_type_count = [0,0]
    for i in range(cols):
        if( column_info_list[i][0] == "STR" ):
            col_type_count[0] += 1
        else:
            col_type_count[1] += 1
    
    if( col_type_count.index( max(col_type_count)) == 0 ):
        print_l("error: data has too many useless columns!")
        print_l("[UnprocessableDataError]")
        Global.exit_flag = True
        return 
    
    print_l(">>>(satisfied)")
    return    

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -   

# None all other type values:
def filter( data, colunm_info_list ):
    
    print_l("\n")
    print_l("*filtering...")
    print_l("*removing other-type values from column")
    print_l(">>>removed values:")
    
    rows = len(data)
    cols = len(data[0])

    count = [ [0] for i in range(cols)]
    
    for col in range(cols):
        for row in range(rows):
            
            val = data[row][col]
            if( val == None):
                continue
            if( colunm_info_list[col][0] == "NUM" ):
               if( type(val) == type("1") ):
                   data[row][col] = None
                   count[col][0] += 1
            elif( type(val) == type(1) or  type(val) == type(1.0) ):
                    data[row][col] = None
                    count[col][0] += 1
    
    for i in range(cols):
        print_l("column " + str(i) + ": [" + str(count[i][0]) + "]")
    
    print_l(">>>done")            
    return data 


#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -   

# converting or removing unnecessary rows and columns in data:
def elimination( data, column_info_list, rows, cols):
    
    global drop_att
    drop_att = []
    
    print_l("\n")
    print_l("*eliminating less informative data...")
    
    THRESHOLD = 30
    MAX_MISS_VAL = 1
    miss_val_threshold = int(rows * THRESHOLD / 100)
    drop_col_list = []
    drop_row_list = []
    
    # dropping column [miss val]
    for col in range(cols):
        if( column_info_list[col][1] > miss_val_threshold ):
            drop_col_list.append(col)
    
    
    
    # if label column to be dropped then return None
    if(cols-1 in drop_col_list):
        print_l("\n")
        print_l("error: data has too many unlabeled records!")
        print_l(">>>[UndefinedDataError]")
        Global.exit_flag = True
        return
    
    
    # checking primary validation for columns
    if( cols - len(drop_col_list) < 2 ):
        print_l("error: dataSet has very less amount of informative data columns!")
        print_l(">>>[DataShortageError]")
        Global.exit_flag = True
        return 
    
    
    # droping column [!numerificable] 
    # separating STR cols
    for col in range(cols):
        
        if( col in drop_col_list ):
            continue
        
        if( column_info_list[col][0] == "STR" ):
            drop_col_list.append(col)
    
    
    # getting drop column index for attributes
    for col in range(cols):
        if(col in drop_col_list):
            drop_att.append(col)
    
    print_l(">>>eliminated column (indices):")
    if( len(drop_col_list) == 0):
        print_l("None")
    else:
        for i in range(len(drop_col_list)):
            print_l( str(drop_col_list[i]) )
    
                    
    # checking primary validation for columns
    if( cols - len(drop_col_list) < 2 ):
        print_l("error: dataSet got very less amount of informative data columns:")
        print_l(">>>[DataShortageError]")
        Global.exit_flag = True
        return       
             
            
    # dropping rows [miss val]  
    for row in range(rows):
        miss_val_count = 0
        
        if(data[row][cols-1] == None): # if label value is missing
            drop_row_list.append(row)
            continue

        for col in range(cols-1):
             
            if(col in drop_col_list):
                continue
            
            if(data[row][col] == None):
                miss_val_count += 1
                
            if(miss_val_count > MAX_MISS_VAL): # if more than 1 values missing
                drop_row_list.append(row)
                break
              
    print_l(">>>eliminated row (indices): ")        
    if( len(drop_row_list) == 0):
        print_l("None")
    else:
        for i in range(len(drop_row_list)):
            print_l( str(drop_row_list[i]) )
    
    
    # checking primary validation for rows
    if( rows - len(drop_row_list) < 10 ):
        print_l("error: data got too less records:")
        print_l(">>>[DataShortageError]")
        Global.exit_flag = True
        return 
                 
    # getting new data list  
    drop_row_count = len(drop_row_list)
    new_data = [[]for i in range(rows - drop_row_count)] # new data list excluding dropping rows
    
    new_data_row = 0 # iterating new data rows separately
    for row in range(rows):
        
        if(row not in drop_row_list):
            
            for col in range(cols):
                
                if(col in drop_col_list):
                    continue
            
                new_data[new_data_row].append(data[row][col])
                
            new_data_row += 1
            
    print_l(">>>done")
    
    print_l("\n")
    print_l("^rows affected:     " + str(drop_row_count))
    print_l("^columns affected:  " + str(len(drop_col_list)))
    
    print_l("\n")
    print_l("dataset dimensions (final):")
    print_l("data columns: [" + str(len(new_data[0])) + "]")
    print_l("data records : [" + str(len(new_data)) + "]")
    
    return new_data
   
   
#=========================================================================================================
# Gui Funcs and callbacks :-

# updating text area:
def print_l(data):
            
    textArea.config(state=NORMAL)
    textArea.insert(END, data+"\n")
    textArea.config(state=DISABLED)
    return

#..........................................................................................................

# thread (t2):
def askForSaveLog():
    while(True):
        time.sleep(0.1)
        if(Global.exit_flag or Global.committed):
            
            # break thread without doing task
            if(Global.exit_flag == True and Global.proceed == False):
                print("t2 breaking loop(1)")
                break
            
            
            if(Global.exit_flag):
                tLabel3.config(bg='#4f0c07')
            else:
                tLabel3.config(bg='#194f0a')
            
            save.place(x=int(0.82685*width), y=int(0.95025*height))
            Global.continue_ = True
            print("t2 breaking loop(2)")
            break

#..........................................................................................................

# auto save:
def auto_save_l():
    
    # if autosave flag is off
    if(not Global.log_auto_save): 
        return
    
    # if no task selected
    if(Global.file_name == None  or  Global.proceed == False ): 
        return
    
    
    # otherwise...
    filename = "/" + "Log_" + Global.file_name + ".txt"
    
    # writing data into the file.
    file = open(Global.folder_path+filename,'w')
    file.write(str(textArea.get("1.0",END))) 
    file.close()
    return

#..........................................................................................................

# close window (label-button):
def c_enter(evnt):
    close.config(bg='#711d1d')

def c_leave(evnt):
    close.config(bg='#1f1e1c')

def c_clicked(event):
    
    if(Global.org_window != None):
        Global.org_window.attributes('-topmost',False)
    if(Global.pro_window != None):
        Global.pro_window.attributes('-topmost',False)
    
    choice = messagebox.askyesno(title="confirm again",message=" Do you want to exit Process?")
    if(choice):
        
        Global.exit_flag = True
        
        if(Global.committed):
            auto_save_l()
            
        elif(Global.proceed):
            print_l("\n")
            print_l("<<<< Process [Closed]")
            auto_save_l()
        
        # Global.exit_flag = True
        time.sleep(0.5)
        window.destroy()
                
    else:
        if(Global.org_window != None):
            Global.org_window.attributes('-topmost',True)
        if(Global.pro_window != None):
            Global.pro_window.attributes('-topmost',True)
        return

#..........................................................................................................

# choose file (label-button):
def f_enter(event):
    file_btn.config(bg='#d9d9d9')

def f_leave(event):
    file_btn.config(bg='#7d7979')

def f_clicked(event):
    
    # placing original data window at backside:
    if(Global.org_window != None):
        Global.org_window.attributes('-topmost',False)
        
    Global.file_path = filedialog.askopenfilename(
        initialdir="C:/Users/Shreyash Patil/Desktop", # intitial directory
        title="Choose a File",
        filetypes=( ("Excel Files","*.xlsx"),("Excel Files","*.xls") )
    )
    
    
   
    if(Global.file_path != ""):
        # if user selected change file and leave without selecting new file
        # then previous file will be processed again
        # hence storing current file path in another veriable i.e. previous_file_path
        Global.previous_file_path = Global.file_path
        # display file data:
        Display.display(0, Global.file_path)
        
        # print("file path: "+Global.file_path)
        Global.folder_path = os.path.dirname(Global.file_path)
        # print("folder path:" +Global.folder_path)
        Global.file_name = Path(Global.file_path).stem
        tLabel1.config(text= "File:   " + Global.file_name, fg='#4fa9c9', bg='#0c2b30', )
        file_btn.config(text="Change")
        go.config(text="Proceed")
        file_btn.place(x=int(0.52085*width),y=int(0.00699*height))
        go.place(x=int(0.60548*width),y=int(0.00699*height))
    
    # if user selected change file and leave without selecting new file
    # then previous file will be processed again
    # hence getting previous file into main file
    if( Global.file_path == ""):
        # if no new file selected:
        Global.file_path = Global.previous_file_path
        
        # showing previous original data window again:
        if(Global.org_window != None):
                Global.org_window.attributes('-topmost',True)
            
    if Global.file_path == "":
        return
    
    
#.......................................................................................................... 

# proceed (label-button):
def go_enter(event):
    go.config(bg='#d9d9d9') 

def go_leave(event):
    go.config(bg='#7d7979')

def go_clicked(event):

    global t1
    global t2
    
    print_l(">>>> Process [started]")
    print_l("\n")
    print_l("[File]: " + Global.file_name)
    print_l("[Path]: " + Global.file_path)
    
    file_btn.place(x=-100,y=-100)
    go.place(x=-100,y=-100)
    Global.proceed = True
    Global.exit_flag = False
    
    # joining previous threads and creating new
    if(Global.continue_):
        t1.join()
        print("t1 joining")
        t2.join()
        print("t2 joining")

        t1 = None
        t2 = None
        
        
        t1 = threading.Thread(target=processor)
        t2 = threading.Thread(target=askForSaveLog)
    
        t1.start()
        print("t1 restarting")
        t2.start()
        print("t2 restarting")
    
        
#..........................................................................................................

# save log (label-button):
def s_clicked(event):
    
    if(Global.org_window != None):
        Global.org_window.attributes('-topmost',False)
    if(Global.pro_window != None):
        Global.pro_window.attributes('-topmost',False)
        
    new_path = filedialog.askdirectory() # getting folder(directory) to save
    
    if new_path == '':
        
        if(Global.org_window != None):
            Global.org_window.attributes('-topmost',True)
        if(Global.pro_window != None):
            Global.pro_window.attributes('-topmost',True)
        return
    
    # log filename
    fileName = "Log_" + Global.file_name + ".txt"  
    
    # merging directory and file name
    location = new_path + "/" + fileName 
    
    # getting data from text area
    data = str(textArea.get("1.0",END)) 
    
    # writing data into the file.
    file = open(location,'w')
    file.write(data) 
    file.close()
    
    
    # reset all for next task
    save.place(x=-100,y=-100)
    textArea.config(state=NORMAL)
    textArea.delete("1.0",END)
    textArea.config(state=DISABLED)
    tLabel3.config(bg='#2f2e2c')
    if(Global.org_window != None):
        Global.org_window.destroy()
        Global.org_window = None
    if(Global.pro_window != None):
        Global.pro_window.destroy()
        Global.pro_window = None
    
    # ask for new task
    time.sleep(1)
    file_btn.place(x=930,y=6)
    Global.proceed = False
    
    
def s_enter(event):
    save.config( bg='#00ff11')

def s_leave(event):
    save.config(bg='#2f2e2c',fg='#000000')


#=========================================================================================================
# task manager func:-
# thread (t1)

def processor():
    
    Global.committed = False
    Global.read_data = False

    while(True):
        time.sleep(1)
        if(Global.proceed == True): # wait until task is assigned
            break
        if(Global.exit_flag == True): # if user quit program by doing nothing then loop must break and join
            return
        
    # get attribute list
    attri_list = getData(False)
    if(Global.exit_flag):
        return
    
    print_l("\n")
    print_l("attributes names: ")
    for i in range(len(attri_list)):
        print_l("column " + str(i) +": [ " + str(attri_list[i]) + " ]")
    
    # get actual data
    data = getData()
    if(Global.exit_flag):
        return
    
    
    # get table dimension
    rows = len(data)
    cols = len(data[0])
    
    # getting column info.
    column_info_list = get_column_info( data, rows, cols )

    # secondary validation
    sec_validation(column_info_list)
    if(Global.exit_flag):
        return
    
    # null other type values from col
    data = filter(data, column_info_list)
    
    # updating missing value list
    print_l("\n\n"+"*updating column information...")
    column_info_list = get_column_info( data, rows, cols )
    print_l(">>>done")
    
    # getting more reliable data by eliminating useless records and columns
    new_data = elimination(data, column_info_list, rows, cols)
    del(data)
    if(Global.exit_flag):
        return
    
    # updating attributes list
    print_l("\n")
    print_l("*updating attribute list...")
    for col in range(len(attri_list)):
        if(col in drop_att):
            del(attri_list[col])

    print_l(">>>done")
    
    # updating data dimention
    rows = len(new_data)
    cols = len(new_data[0])
    
    # updating column information
    print_l("\n")
    print_l(">>>updated column informtion:")
    column_info_list = get_column_info( new_data, rows, cols )
    if(Global.exit_flag):
        return
    
    # checking for missing value
    print_l("\n")
    print_l("*checking for missing values in new data...")

    has_missing_val = False
    for col in range(cols):
        if(column_info_list[col][1] > 0):
            
            pro_data = None
            has_missing_val = True
            
            print_l(">>>missing values detected")
            print_l("\n")
            print_l("*filling missing values...")
            
            pro_data = setMissing(new_data, rows, cols) 
            
            if(Global.exit_flag):
                print_l("Data got too many missing values!...Can`t Process Further")
                print_l("[BadDataError!]")
                return
            
            print_l(">>>done") 
            
            # getting column info.
            column_info_list = get_column_info( new_data, rows, cols )

            new_data = pro_data
            del(pro_data)    
            break
    
     
    if(not has_missing_val):
        print_l(">>>None")
    
    
    # exporting processed data
    print_l("\n")
    print_l("*exporting processed data...")
    writeExcel(attri_list, new_data)
    print_l(">>>data exported successfully!")
    Global.pro_file_path = Global.folder_path + "/pro_" +  Global.file_name + ".xlsx"
    print_l("@Location:  " + Global.pro_file_path)

    # display processed data
    Display.display(1, Global.pro_file_path)

    print_l("\n")
    print_l("<<<<Process [Committed]") 
    Global.committed = True
        
    
#=========================================================================================================
# GUI definition:-

# thread t1 starts
t1 = threading.Thread(target=processor)
t1.start()
print("t1 started")

# thread t2 starts
t2 = threading.Thread(target=askForSaveLog)
t2.start()
print("t2 started")

# gui windows mainloop:
window = Tk()
    
window.config(bg='#000000')
window.overrideredirect(TRUE) 

width = window.winfo_screenwidth()
height = window.winfo_screenheight()

window.geometry("%dx%d" % (width, height))

titleBar = Label(window,text="DataFilter",font=('Consolas',15),bg='#1f1e1c',fg='#4f4f4f',height=int(0.0012*height), pady=int(0.00116*height),width=int(0.0912*width),  anchor=SW)
titleBar.place(x=0,y=0)

close = Label(window,  text="X",  padx=int(0.0046*width), font=('Roboto',15), bg='#1f1e1c', fg='#000000') 
close.place(x=int(width/2), y=0)

close.bind("<Button-1>", c_clicked)
close.bind("<Enter>", c_enter)
close.bind("<Leave>", c_leave)

textArea = Text(window, bg="#0c0f14", font=("Consloas", 14), height=int(0.0371*height), width=int(0.0261*width), padx=int(0.00977*width), pady=int(0.0232*height), fg = "green", border=NO, wrap='word',state=DISABLED)
textArea.place(x= int((0.1954*width) + width/2), y=int(0.0822*height))

tLabel1 = Label(window, bg='#300c0d',text="No File Selected" ,fg='#ff0000', height=int(0.00232*height), width=int(0.0866*width), font=('Times New Roman CYR', 10))
tLabel1.place(x=int(-0.00265*width), y=int(0.0359*height))

file_btn = Label(tLabel1, bg='#7d7979', text="Chose File",font=('Times New Roman CYR', 10), height=int(0.001158*height),width=int(0.0098*width),padx=int(0.00131*width),pady=int(0.001156*height) )
file_btn.place(x=int(0.60548*width),y=int(0.00699*height))
file_btn.bind("<Button-1>", f_clicked)
file_btn.bind("<Enter>", f_enter)
file_btn.bind("<Leave>", f_leave)

go = Label(tLabel1, bg='#7d7979', text="Proceed",font=('Times New Roman CYR', 10), height=int(0.001158*height),width=int(0.0098*width),padx=int(0.00131*width),pady=int(0.001156*height) )
go.bind("<Button-1>", go_clicked)
go.bind("<Enter>", go_enter)
go.bind("<Leave>", go_leave)

tLabel2 = Label(window, text="L o g   I n f o r m a t i o n", font=("Consloas", 10), bg='#2f2e2c', fg='#4ecf23', height=int(0.002316*height), width=int(0.0378*width))
tLabel2.place(x= int((0.1954*width) + width/2), y=int(0.0359*height))

tLabel3 = Label(window, font=("Consloas", 10), bg='#2f2e2c', height=int(0.002316*height), width=int(0.0378*width))
tLabel3.place(x= int((0.1954*width) + width/2), y=int(0.9457*height))

save = Label(window, bg='#2f2e2c',font=('Times New Roman CYR', 10), fg='#000000' , text="SAVE", height=int(0.001158*height), width=int(0.00656/width), padx=int(0.00131*width), pady=int(0.0058*height))
save.bind("<Button-1>", s_clicked)
save.bind("<Enter>", s_enter)
save.bind("<Leave>", s_leave)

window.mainloop()   
        
        
# threads joining at the end of program
t1.join()
print("t1 joined at end")
t2.join()
print("t2 joined at end")